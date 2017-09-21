import cv2
from string import uppercase as AtoZ
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ????
image = cv2.imread("test.jpg")

# ??Excel
wb = load_workbook(filename='test.xlsx')

# ???sheet??
ws = wb.active

# ??????
# col_name = ['A', 'B', ...., 'Z']
col_name = list(AtoZ)

# excel????
for i in xrange(len(image[0]) - 26):
    col_name.append(AtoZ[i // 26] + AtoZ[i % 26])

for col in xrange(len(image[0])):
    # ??????
    ws.column_dimensions[col_name[col]].width = 0.3

for row in xrange(len(image)):
    # ??????
    ws.row_dimensions[row].height = 1.5

    for col in xrange(len(image[row])):
        # ????????16??,?????????
        red, green, blue = image[row][col]
        color = '%02x%02x%02x' % (blue, green, red)

        # ??????????
        cell_name = col_name[col] + str(row + 1)
        # ??????
        cell = ws[cell_name]
        # ????
        cell.fill = PatternFill(patternType='solid', fgColor=color)

# excel??
wb.save('test.xlsx')